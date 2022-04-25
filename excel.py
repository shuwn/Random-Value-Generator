# -*- coding: utf-8 -*-
from importlib.resources import path
import os, sys, subprocess
from createnewwindow import CreateNewWindow
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from pathlib import Path

def open_file(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])
        
class Table:
    def __init__(self, Project_Name, database, items):
        if len(Project_Name) == 0:
            CreateNewWindow.warningWin("Please enter Project Name.")
        else:
        # create Excel
            excel = Workbook()
            datasheet = excel.active
            datasheet.title = str(Project_Name)

            # setting table Width
            datasheet.column_dimensions['A'].width = 10
            datasheet.column_dimensions['B'].width = 20

            # setting table Tielte
            datasheet.merge_cells('A1:B1')
            datasheet['A1'] = Project_Name
            module_Name_Align = Alignment(horizontal='center', vertical='center')
            module_Name_Font = Font(name='Calibri', color='000000', size=16, b=True)
            datasheet['A1'].alignment = module_Name_Align
            datasheet['A1'].font = module_Name_Font

            # Title
            LV_font = Font(name='Calibri', color='f1f1f1', size=14, b=True)
            LV_fill = PatternFill("solid", fgColor="165896")
            LV_Align = Alignment(horizontal='center', vertical='center')
            datasheet.merge_cells('A2:B2')
            datasheet['A2'] = 'Statistics'
            datasheet['A2'].font = LV_font
            datasheet['A2'].fill = LV_fill
            datasheet['A2'].alignment = LV_Align

            # Calculate the maximum, minimum and average value
            datasheet['A3'] = 'Max.'
            datasheet['A4'] = 'Min.'
            datasheet['A5'] = 'Avg.'
            datasheet['B3'].value = round(max(database), 3)
            datasheet['B4'].value = round(min(database), 3)
            datasheet['B5'].value = round(sum(database)/items, 3)

            # Set up iteam and fill in the content
            for d in range(0, items):
                datasheet.cell(row=d+6, column=1).value = d + 1
                datasheet.cell(row=d+6, column=2).value = database[d]

            # setting content style
            title_Align = Alignment(horizontal='center', vertical='center')
            title_Font = Font(name='Calibri', color='000000', size=14, b=True)
            content_Font = Font(name='Calibri', color='000000', size=14)

            for f in range(0, items+3):
                datasheet.cell(row=f + 3, column=1).alignment = title_Align
                datasheet.cell(row=f + 3, column=1).font = title_Font
                datasheet.cell(row=f + 3, column=2).alignment = title_Align
                datasheet.cell(row=f + 3, column=2).font = content_Font

            max_Rows = datasheet.max_row


            # draw grid
            side = Side(style = 'thin', color='000000')

            for rows in datasheet['A2':'B' + str(max_Rows)]:
                for cell in rows:
                    cell.border = Border(left = side, right = side, top = side, bottom = side)

            Table.set_Border(datasheet, "A2:" + "B" +str(len(datasheet['B'])))
            Table.set_Border(datasheet, "A3:B5")
            # Seve Excel
            excel_filename = Project_Name + '.xlsx'
            p = Path.home() / 'Downloads'
            root = p / excel_filename
            excel.save(root)
            path = str(root)
            CreateNewWindow.SaveWin(path)
            

    # define outline
    def set_Border(ws, cell_range):
        rows = ws[cell_range]
        side2 = Side(style = 'thick', color='000000')
        rows = list(rows) # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1 # index of the last cell
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1 # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side2
                if pos_x == max_x:
                    border.right = side2
                if pos_y == 0:
                    border.top = side2
                if pos_y == max_y:
                    border.bottom = side2

                # 設置當邊緣格是邊緣時才設置邊框
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border
